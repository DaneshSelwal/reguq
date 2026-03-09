"""Export helpers for Excel, JSON, and plots."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

from .config import coerce_output_config
from .types import ConformalResult, OutputConfig, PhaseResult, TuningResult


def make_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def resolve_output_dir(output_config: OutputConfig | dict[str, Any] | None) -> Path | None:
    cfg = coerce_output_config(output_config)
    if cfg.output_dir is None:
        return None
    output_dir = Path(cfg.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def safe_sheet_name(name: str) -> str:
    cleaned = name.replace("/", "_").replace("\\", "_").replace(" ", "_")
    return cleaned[:31]


# Backward compatible internal alias used by earlier code.
_safe_sheet_name = safe_sheet_name


def _excel_engine() -> str:
    try:
        import xlsxwriter  # noqa: F401

        return "xlsxwriter"
    except Exception:
        return "openpyxl"


def write_phase_excel(result: PhaseResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine=_excel_engine()) as writer:
        result.metrics.to_excel(writer, sheet_name="metrics", index=False)
        if result.params:
            params_rows = [
                {"model": model, "params": json.dumps(params, sort_keys=True)}
                for model, params in result.params.items()
            ]
            pd.DataFrame(params_rows).to_excel(writer, sheet_name="params", index=False)
        for model_id, pred_df in result.predictions.items():
            sheet = safe_sheet_name(f"pred_{model_id}")
            pred_df.to_excel(writer, sheet_name=sheet, index=False)
    return output_path


def write_tuning_excel(result: TuningResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine=_excel_engine()) as writer:
        result.summary.to_excel(writer, sheet_name="summary", index=False)
        for model_id, pred_df in result.predictions.items():
            sheet = safe_sheet_name(f"pred_{model_id}")
            pred_df.to_excel(writer, sheet_name=sheet, index=False)
    return output_path


def write_conformal_excel(result: ConformalResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine=_excel_engine()) as writer:
        for method_name, method_result in result.methods.items():
            method_result.metrics.to_excel(
                writer,
                sheet_name=safe_sheet_name(f"m_{method_name}"),
                index=False,
            )
            for model_id, pred_df in method_result.predictions.items():
                sheet = safe_sheet_name(f"{method_name}_{model_id}")
                pred_df.to_excel(writer, sheet_name=sheet, index=False)
    return output_path


def write_json(data: Any, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")
    return output_path


def save_interval_plot(
    pred_df: pd.DataFrame,
    output_path: Path,
    title: str,
    max_points: int = 300,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    subset = pred_df.head(max_points)
    x = range(len(subset))
    plt.figure(figsize=(10, 4))
    plt.plot(x, subset["y_true"], label="y_true", linewidth=1.5)
    if "y_pred" in subset:
        plt.plot(x, subset["y_pred"], label="y_pred", linewidth=1.2)
    if "y_lower" in subset and "y_upper" in subset:
        plt.fill_between(x, subset["y_lower"], subset["y_upper"], alpha=0.2, label="interval")
    plt.title(title)
    plt.legend(loc="best")
    plt.tight_layout()
    plt.savefig(output_path, dpi=160)
    plt.close()
    return output_path


def _anchor_for_index(start_cell: str, row_step: int, index: int) -> str:
    col_letter, row = coordinate_from_string(start_cell)
    col_idx = column_index_from_string(col_letter)
    target_row = row + index * row_step
    return f"{get_column_letter(col_idx)}{target_row}"


def embed_images_in_excel(
    workbook_path: Path,
    images_by_sheet: dict[str, list[Path]],
    start_cell: str = "J2",
    row_step: int = 22,
) -> Path:
    if not images_by_sheet:
        return workbook_path

    wb = load_workbook(workbook_path)
    for sheet_name, image_paths in images_by_sheet.items():
        if not image_paths:
            continue
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=safe_sheet_name(sheet_name))
        ws = wb[safe_sheet_name(sheet_name)]

        img_index = 0
        for image_path in image_paths:
            if not image_path.exists():
                continue
            anchor = _anchor_for_index(start_cell=start_cell, row_step=row_step, index=img_index)
            ws.add_image(OpenpyxlImage(str(image_path)), anchor)
            img_index += 1

    wb.save(workbook_path)
    return workbook_path
