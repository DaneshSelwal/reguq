from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from reguq.quantile import run_quantile


def test_quantile_excel_export_schema(tmp_path: Path, synthetic_paths, patched_estimators):
    train_path, test_path = synthetic_paths
    output_dir = tmp_path / "outputs"

    result = run_quantile(
        data={"train_path": train_path, "test_path": test_path},
        target_col="target",
        models=["lightgbm", "xgboost"],
        params_source={"mode": "defaults"},
        output_config={
            "output_dir": output_dir,
            "export_excel": True,
            "export_plots": True,
            "embed_excel_charts": True,
            "save_json": True,
        },
    )

    excel_path = output_dir / "quantile.xlsx"
    assert excel_path.exists()
    wb = load_workbook(excel_path)
    assert "metrics" in wb.sheetnames
    assert "pred_lightgbm" in wb.sheetnames
    assert "pred_xgboost" in wb.sheetnames
    assert len(wb["pred_lightgbm"]._images) >= 1
    assert len(wb["metrics"]._images) >= 1

    assert any(p.suffix == ".xlsx" for p in result.artifacts)
    assert any(p.suffix == ".png" for p in result.artifacts)


def test_conformal_excel_export_schema(tmp_path: Path, synthetic_paths, patched_estimators):
    train_path, test_path = synthetic_paths
    output_dir = tmp_path / "outputs_conformal"
    
    from reguq.conformal_standard import run_conformal_standard
    result = run_conformal_standard(
        data={"train_path": train_path, "test_path": test_path},
        target_col="target",
        models=["lightgbm"],
        params_source={"mode": "defaults"},
        conformal_config={
            "alpha": 0.1,
            "methods": ["mapie"],
        },
        output_config={
            "output_dir": output_dir,
            "export_excel": True,
            "export_plots": True,
            "embed_excel_charts": True,
        }
    )
    
    excel_path = output_dir / "conformal_standard.xlsx"
    assert excel_path.exists()
    
    wb = load_workbook(excel_path)
    # Check that sheets are combined
    assert "metrics" in wb.sheetnames
    assert "predictions" in wb.sheetnames
    
    # Read the data back using pandas to check columns
    pred_df = pd.read_excel(excel_path, sheet_name="predictions")
    expected_cols = ["sample_index", "strategy", "model", "y_true", "y_pred", "ymin", "ymax", "width", "is_covered", "residual", "backend"]
    for col in expected_cols:
        assert col in pred_df.columns
    # Check that y_lower and y_upper are NOT in the Excel predictions
    assert "y_lower" not in pred_df.columns
    assert "y_upper" not in pred_df.columns

